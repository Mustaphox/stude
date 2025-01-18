from flask import Flask, render_template, request, redirect, url_for, Response, flash, send_file, session
import sqlite3
import os
import io
import re
import xlsxwriter
import tkinter as tk
import openpyxl
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'default_secret_key')

# Function to connect to the database
def get_db_connection():
    conn = sqlite3.connect('school.db')
    conn.row_factory = sqlite3.Row
    return conn

# Initialize database tables if they do not exist
def initialize_database():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            specialty TEXT NOT NULL,
            level TEXT NOT NULL,
            year TEXT NOT NULL,
            user_id INTEGER NOT NULL,  -- Ensure this line is present
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            class_id INTEGER NOT NULL,
            FOREIGN KEY (class_id) REFERENCES classes (id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            surname TEXT NOT NULL,
            sessions_attended INTEGER DEFAULT 0,
            group_id INTEGER NOT NULL,
            FOREIGN KEY (group_id) REFERENCES groups (id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            date DATE NOT NULL,
            time TIME NOT NULL,
            FOREIGN KEY (group_id) REFERENCES groups (id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            student_id INTEGER,
            session_id INTEGER,
            status TEXT, -- 'present', 'absent', or 'justified'
            observation TEXT,
            PRIMARY KEY (student_id, session_id),
            FOREIGN KEY (student_id) REFERENCES students(id),
            FOREIGN KEY (session_id) REFERENCES sessions(id)
        )
    ''')
    conn.commit()
    conn.close()

# Initialize the database when the app starts
initialize_database()

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Route for the login page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            flash('Logged in successfully!', 'success')
            return redirect(url_for('welcome'))
        else:
            flash('Invalid username or password.', 'error')

    return render_template('login.html')

# Route for user registration (signup)
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if not username or not password:
            flash('Username and password are required.', 'error')
            return redirect(url_for('signup'))

        hashed_password = generate_password_hash(password)

        conn = get_db_connection()
        try:
            conn.execute('INSERT INTO users (username, password) VALUES (?, ?)', (username, hashed_password))
            conn.commit()
            flash('Account created successfully! Please log in.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username already exists. Please choose a different one.', 'error')
        finally:
            conn.close()

    return render_template('signup.html')

# Route for the welcome page
@app.route('/welcome')
@login_required
def welcome():
    conn = get_db_connection()
    try:
        # Fetch data only for the logged-in user
        total_classes = conn.execute(
            'SELECT COUNT(*) FROM classes WHERE user_id = ?', (session['user_id'],)
        ).fetchone()[0]

        total_groups = conn.execute(
            'SELECT COUNT(*) FROM groups WHERE class_id IN (SELECT id FROM classes WHERE user_id = ?)',
            (session['user_id'],)
        ).fetchone()[0]

        total_students = conn.execute(
            'SELECT COUNT(*) FROM students WHERE group_id IN (SELECT id FROM groups WHERE class_id IN (SELECT id FROM classes WHERE user_id = ?))',
            (session['user_id'],)
        ).fetchone()[0]

        return render_template(
            'index.html',
            total_classes=total_classes,
            total_groups=total_groups,
            total_students=total_students
        )
    except Exception as e:
        return f"An error occurred: {e}"
    finally:
        conn.close()

# Route for logging out
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))

# Route to display all classes
@app.route('/classes')
@login_required
def classes():
    conn = get_db_connection()
    classes = conn.execute('SELECT * FROM classes WHERE user_id = ?', (session['user_id'],)).fetchall()
    conn.close()
    return render_template('classes.html', classes=classes)

# Function to validate the year range format
def validate_year_range(year):
    if re.match(r'^\d{4}-\d{4}$', year):
        return True
    return False

# Route to add a class
@app.route('/add-class', methods=['GET', 'POST'])
@login_required
def add_class():
    if request.method == 'POST':
        name = request.form['name']
        specialty = request.form['specialty']
        level = request.form['level']
        year = request.form['year']

        if not validate_year_range(year):
            flash('Invalid year range format. Please use the format YYYY-YYYY (e.g., 2024-2025).', 'error')
            return redirect(url_for('add_class'))

        conn = get_db_connection()
        conn.execute(
            'INSERT INTO classes (name, specialty, level, year, user_id) VALUES (?, ?, ?, ?, ?)',
            (name, specialty, level, year, session['user_id'])
        )
        conn.commit()
        conn.close()
        flash('Class added successfully!', 'success')
        return redirect(url_for('classes'))

    return render_template('add-class.html')

# Route to edit a class
@app.route('/edit-class/<int:class_id>', methods=['GET', 'POST'])
@login_required
def edit_class(class_id):
    conn = get_db_connection()
    class_data = conn.execute('SELECT * FROM classes WHERE id = ? AND user_id = ?', (class_id, session['user_id'])).fetchone()

    if request.method == 'POST':
        name = request.form['name']
        specialty = request.form['specialty']
        level = request.form['level']
        year = request.form['year']

        if not validate_year_range(year):
            flash('Invalid year range format. Please use the format YYYY-YYYY (e.g., 2024-2025).', 'error')
            return redirect(url_for('edit_class', class_id=class_id))

        conn.execute(
            'UPDATE classes SET name = ?, specialty = ?, level = ?, year = ? WHERE id = ?',
            (name, specialty, level, year, class_id)
        )
        conn.commit()
        conn.close()
        flash('Class updated successfully!', 'success')
        return redirect(url_for('classes'))

    conn.close()
    return render_template('edit-class.html', class_data=class_data)

# Route to delete a class
@app.route('/delete-class/<int:class_id>', methods=['POST'])
@login_required
def delete_class(class_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM classes WHERE id = ? AND user_id = ?', (class_id, session['user_id']))
    conn.commit()
    conn.close()
    return redirect('/classes')

# Main route
@app.route('/')
def index():
    return redirect('/login')
# Route to display groups for a specific class
@app.route('/class/<int:class_id>/groups')
@login_required
def groups(class_id):
    conn = get_db_connection()
    class_data = conn.execute('SELECT * FROM classes WHERE id = ? AND user_id = ?', (class_id, session['user_id'])).fetchone()
    groups = conn.execute('SELECT * FROM groups WHERE class_id = ?', (class_id,)).fetchall()
    conn.close()
    return render_template('groups.html', class_data=class_data, groups=groups, class_id=class_id)

# Route to add a group to a class
@app.route('/class/<int:class_id>/add-group', methods=['GET', 'POST'])
@login_required
def add_group(class_id):
    if request.method == 'POST':
        group_type = request.form['group-type']
        conn = get_db_connection()
        if group_type == "TP/TD":
            conn.execute('INSERT INTO groups (type, class_id) VALUES (?, ?)', ("TD", class_id))
            conn.execute('INSERT INTO groups (type, class_id) VALUES (?, ?)', ("TP", class_id))
        else:
            conn.execute('INSERT INTO groups (type, class_id) VALUES (?, ?)', (group_type, class_id))
        conn.commit()
        conn.close()
        return redirect(url_for('groups', class_id=class_id))

    return render_template('add-group.html', class_id=class_id)

# Route to edit a group
@app.route('/edit-group/<int:group_id>', methods=['GET', 'POST'])
@login_required
def edit_group(group_id):
    conn = get_db_connection()
    group_data = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()

    if request.method == 'POST':
        group_type = request.form['group-type']
        conn.execute('UPDATE groups SET type = ? WHERE id = ?', (group_type, group_id))
        conn.commit()
        conn.close()
        return redirect(url_for('groups', class_id=group_data['class_id']))

    conn.close()
    return render_template('edit-group.html', group_data=group_data)

# Route to delete a group
@app.route('/delete-group/<int:group_id>', methods=['POST'])
@login_required
def delete_group(group_id):
    conn = get_db_connection()
    group_data = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()
    conn.execute('DELETE FROM groups WHERE id = ?', (group_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('groups', class_id=group_data['class_id']))

# Route to export attendance
@app.route('/export-attendance/<int:group_id>', methods=['GET', 'POST'])
@login_required
def export_attendance(group_id):
    if request.method == 'POST':
        date_debut = request.form.get('date_debut')
        date_fin = request.form.get('date_fin')

        if not date_debut or not date_fin:
            return "Please check the date range.", 400

        conn = get_db_connection()
        sessions = conn.execute('''
            SELECT id, date 
            FROM sessions 
            WHERE group_id = ? AND date BETWEEN ? AND ?
            ORDER BY date
        ''', (group_id, date_debut, date_fin)).fetchall()

        if not sessions:
            conn.close()
            return "No sessions or wrong date.", 404

        students = conn.execute('''
            SELECT id, name, surname 
            FROM students 
            WHERE group_id = ?
            ORDER BY id
        ''', (group_id,)).fetchall()

        session_ids = tuple(session['id'] for session in sessions)
        attendance_data = conn.execute(f'''
            SELECT 
                attendance.student_id,
                attendance.session_id,
                attendance.status
            FROM attendance
            WHERE attendance.session_id IN ({','.join('?' * len(session_ids))})
        ''', session_ids).fetchall()
        conn.close()

        attendance_dict = {}
        for record in attendance_data:
            student_id = record['student_id']
            session_id = record['session_id']
            attendance_dict[(student_id, session_id)] = record['status']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Group {group_id} Attendance"

        headers = ['Student Name', 'Student Surname']
        headers.extend([f" ({session['date']})" for session in sessions])
        ws.append(headers)

        for student in students:
            row = [student['name'], student['surname']]
            for session in sessions:
                status = attendance_dict.get((student['id'], session['id']), 'Present') 
                row.append(status)
            ws.append(row)
        filepath = os.path.join(os.getcwd(), f"group_{group_id}_attendance_report.xlsx")       
        wb.save(filepath)

        return send_file(filepath, as_attachment=True, download_name=f'group_{group_id}_attendance_report.xlsx')

    return render_template('export-attendance.html')

# Route to view students in a group
@app.route('/group/<int:group_id>/students')
@login_required
def view_students(group_id):
    try:
        conn = get_db_connection()
        group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()
        if not group:
            conn.close()
            return f"Group with ID {group_id} not found.", 404

        students = conn.execute('''
            SELECT s.id, s.name, s.surname, 
                   (SELECT COUNT(*) FROM attendance WHERE attendance.student_id = s.id AND attendance.status = 'present') AS sessions_attended
            FROM students s
            WHERE s.group_id = ?
        ''', (group_id,)).fetchall()

        total_students = len(students)
        conn.close()

        return render_template('students.html', group=group, students=students, total_students=total_students)
    except Exception as e:
        return f"An error occurred: {str(e)}"

# Route to add a student to a group
@app.route('/group/<int:group_id>/student/new', methods=['GET', 'POST'])
@login_required
def add_student(group_id):
    conn = get_db_connection()
    group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()

    if not group:
        conn.close()
        return f"Group with ID {group_id} not found.", 404

    if request.method == 'POST':
        name = request.form.get('name')
        surname = request.form.get('surname')

        if not name or not surname:
            conn.close()
            return "Name and surname are required.", 400

        conn.execute(
            'INSERT INTO students (name, surname, group_id) VALUES (?, ?, ?)',
            (name, surname, group_id)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('view_students', group_id=group_id))

    conn.close()
    return render_template('add-student.html', group=group)

# Route to add students from an Excel file
@app.route('/add_students_excel/<int:group_id>', methods=['POST'])
@login_required
def add_students_excel(group_id):
    if 'excel_file' not in request.files:
        return "No file part", 400

    file = request.files['excel_file']
    if file.filename == '':
        return "No selected file", 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return "Invalid file format. Please upload an Excel file.", 400

    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    uploaded_students = []
    with get_db_connection() as conn:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, surname = row
            if name and surname:
                conn.execute(
                    'INSERT INTO students (name, surname, group_id) VALUES (?, ?, ?)',
                    (name, surname, group_id)
                )
                uploaded_students.append({'name': name, 'surname': surname})
        conn.commit()

    with get_db_connection() as conn:
        group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()
    return render_template('add-student.html', group=group, uploaded_students=uploaded_students)

# Route to export students to an Excel file
@app.route('/export_students/<int:group_id>')
@login_required
def export_students(group_id):
    conn = get_db_connection()
    students = conn.execute(
        'SELECT name, surname FROM students WHERE group_id = ?', (group_id,)
    ).fetchall()
    conn.close()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    sheet.append(["Name", "Surname"])
    for student in students:
        sheet.append([student["name"], student["surname"]])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=students_group_{group_id}.xlsx"}
    )

# Route to edit a student
@app.route('/edit_student/<int:student_id>/<int:group_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id, group_id):
    conn = get_db_connection()
    student = conn.execute('SELECT * FROM students WHERE id = ?', (student_id,)).fetchone()

    if not student:
        conn.close()
        flash('Student not found.', 'error')
        return redirect(url_for('view_students', group_id=group_id))
    
    if request.method == 'POST':
        name = request.form.get('name')
        surname = request.form.get('surname')

        if not name or not surname:
            flash('Name and surname are required.', 'error')
        else:
            try:
                conn.execute(
                    'UPDATE students SET name = ?, surname = ? WHERE id = ?',
                    (name, surname, student_id)
                )
                conn.commit()
                flash('Student updated successfully!', 'success')
            except Exception as e:
                flash(f'Error updating student: {str(e)}', 'error')
        
        conn.close()
        return redirect(url_for('view_students', group_id=group_id))
    
    conn.close()
    return render_template('edit-student.html', student=student, group_id=group_id)

# Route to delete a student
@app.route('/group/<int:group_id>/student/<int:student_id>/delete', methods=['GET'])
@login_required
def delete_student(group_id, student_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM students WHERE id = ?', (student_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('view_students', group_id=group_id))

# Route to delete all students in a group
@app.route('/group/<int:group_id>/student/delete', methods=['GET'])
@login_required
def delete_students(group_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM students WHERE group_id = ?', (group_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('view_students', group_id=group_id))

# Route to manage students in a session
@app.route('/session/<int:session_id>/manage_students', methods=['GET', 'POST'])
@login_required
def manage_students(session_id):
    conn = get_db_connection()
    session_data = conn.execute('SELECT id, date, time FROM sessions WHERE id = ?', (session_id,)).fetchone()

    if not session_data:
        conn.close()
        return "Session not found", 404

    group_id = conn.execute('SELECT group_id FROM sessions WHERE id = ?', (session_id,)).fetchone()['group_id']

    students = conn.execute('''
        SELECT s.id, s.name, s.surname, a.status, a.observation
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id AND a.session_id = ?
        WHERE s.group_id = ?
    ''', (session_id, group_id)).fetchall()

    conn.close()
    return render_template('manage_student.html', group_id=group_id, session_id=session_id, students=students)

# Route to save attendance
@app.route('/save_attendance/<group_id>/<session_id>', methods=['POST'])
@login_required
def save_attendance(group_id, session_id):
    conn = get_db_connection()
    
    try:
        students = conn.execute('SELECT * FROM students WHERE group_id = ?', (group_id,)).fetchall()

        for student in students:
            student_id = student['id']
            status = request.form.get(f'attendance_{student_id}[status]')
            observation = request.form.get(f'attendance_{student_id}[observation]')

            existing_attendance = conn.execute('SELECT * FROM attendance WHERE student_id = ? AND session_id = ?',
                                               (student_id, session_id)).fetchone()

            if existing_attendance:
                conn.execute('''UPDATE attendance 
                                SET status = ?, observation = ? 
                                WHERE student_id = ? AND session_id = ?''',
                             (status, observation, student_id, session_id))
            else:
                conn.execute('''INSERT INTO attendance (student_id, session_id, status, observation) 
                                VALUES (?, ?, ?, ?)''',
                             (student_id, session_id, status, observation))

            if status == 'present':
                conn.execute('UPDATE students SET sessions_attended = sessions_attended + 1 WHERE id = ?',
                             (student_id,))

        conn.commit()
        return redirect(url_for('view_sessions', group_id=group_id))
    except Exception as e:
        conn.rollback()
        return f"An error occurred: {str(e)}", 500
    finally:
        conn.close()

# Route to view sessions for a group
@app.route('/group/<int:group_id>/sessions', methods=['GET'])
@login_required
def view_sessions(group_id):
    conn = get_db_connection()
    try:
        sessions = conn.execute('SELECT * FROM sessions WHERE group_id = ?', (group_id,)).fetchall()
        group_info = conn.execute('SELECT class_id FROM groups WHERE id = ?', (group_id,)).fetchone()
        if not group_info:
            return "Group not found", 404

        class_id = group_info['class_id']

        session_data = []
        for session in sessions:
            session_id = session['id']
            present_count = conn.execute('''
                SELECT COUNT(*) 
                FROM attendance 
                WHERE session_id = ? AND status = 'present'
            ''', (session_id,)).fetchone()[0]

            absent_count = conn.execute('''
                SELECT COUNT(*) 
                FROM attendance 
                WHERE session_id = ? AND status = 'absent'
            ''', (session_id,)).fetchone()[0]

            session_data.append({
                'id': session['id'],
                'date': session['date'],
                'time': session['time'],
                'present_count': present_count,
                'absent_count': absent_count
            })

        return render_template('sessions.html', sessions=session_data, group_id=group_id, class_id=class_id)
    except Exception as e:
        return f"An error occurred: {e}"
    finally:
        conn.close()

# Route to add a session
@app.route('/group/<int:group_id>/session/add', methods=['GET', 'POST'])
@login_required
def add_session(group_id):
    conn = get_db_connection()
    group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()
    if not group:
        return "Group not found", 404

    class_id = group['class_id']

    if request.method == 'POST':
        session_date = request.form['session-date']
        session_time = request.form['session-time']

        conn.execute(
            'INSERT INTO sessions (group_id, date, time) VALUES (?, ?, ?)',
            (group_id, session_date, session_time)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('view_sessions', group_id=group_id))

    conn.close()
    return render_template('add-session.html', group_id=group_id, class_id=class_id)

# Route to edit a session
@app.route('/group/<int:group_id>/session/edit/<int:session_id>', methods=['GET', 'POST'])
@login_required
def edit_session(group_id, session_id):
    conn = get_db_connection()
    session = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()

    if not session:
        conn.close()
        return "Session not found", 404

    if request.method == 'POST':
        session_date = request.form['session-date']
        session_time = request.form['session-time']

        try:
            conn.execute(
                'UPDATE sessions SET date = ?, time = ? WHERE id = ?',
                (session_date, session_time, session_id)
            )
            conn.commit()
            return redirect(url_for('view_sessions', group_id=group_id))
        except Exception as e:
            return f"An error occurred while updating the session: {e}"
        finally:
            conn.close()

    conn.close()
    return render_template('edit-session.html', session=session, group_id=group_id)

# Route to delete a session
@app.route('/group/<int:group_id>/session/delete/<int:session_id>', methods=['POST'])
@login_required
def delete_session(group_id, session_id):
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM sessions WHERE id = ?', (session_id,))
        conn.commit()
        return redirect(url_for('view_sessions', group_id=group_id))
    except Exception as e:
        return f"An error occurred while deleting the session: {e}"
    finally:
        conn.close()

# Route to export session attendance
@app.route('/export_session/<int:session_id>')
@login_required
def export_session(session_id):
    conn = get_db_connection()
    session = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if not session:
        return "Session not found", 404

    students = conn.execute('''
        SELECT s.name, s.surname, a.status, a.observation
        FROM students s
        JOIN attendance a ON s.id = a.student_id
        WHERE a.session_id = ?
    ''', (session_id,)).fetchall()

    conn.close()

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Session Attendance')

    worksheet.write('A1', 'Session ID')
    worksheet.write('B1', session['id'])
    worksheet.write('A2', 'Group ID')
    worksheet.write('B2', session['group_id'] if 'group_id' in session.keys() else 'N/A')
    worksheet.write('A3', 'Name')
    worksheet.write('B3', session['name'] if 'name' in session.keys() else 'N/A')
    worksheet.write('A4', 'Date')
    worksheet.write('B4', session['date'] if 'date' in session.keys() else 'N/A')
    worksheet.write('A5', 'Time')
    worksheet.write('B5', session['time'] if 'time' in session.keys() else 'N/A')

    worksheet.write('A7', 'Name')
    worksheet.write('B7', 'Surname')
    worksheet.write('C7', 'Status')
    worksheet.write('D7', 'Observation')

    row = 7
    for student in students:
        worksheet.write(row, 0, student['name'])
        worksheet.write(row, 1, student['surname'])
        worksheet.write(row, 2, student['status'])
        worksheet.write(row, 3, student['observation'])
        row += 1 

    workbook.close()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'session_{session_id}_attendance.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
        app.run(debug=True, host='0.0.0.0', port=5000)
