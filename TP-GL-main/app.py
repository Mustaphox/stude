from flask import Flask, render_template, request, redirect, url_for, Response, flash, send_file
import sqlite3
import os
import keyboard
import webview
import io 
import re
import xlsxwriter
import tkinter as tk # nhotou toule wl3orthe fl applicationwebimport io
#import keyboard # min dir f11
import openpyxl
from io import BytesIO

app = Flask(__name__)
# Use an environment variable to set the secret key
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'default_secret_key')

# Function to connect to the database
def get_db_connection():
    conn = sqlite3.connect('school.db')
    conn.row_factory = sqlite3.Row
    return conn

# Initialize database tables if they do not exist
def initialize_database():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            specialty TEXT NOT NULL,
            level TEXT NOT NULL,
            year TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            class_id INTEGER NOT NULL,
            FOREIGN KEY (class_id) REFERENCES classes (id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            surname TEXT NOT NULL,
            sessions_attended INTEGER DEFAULT 0,
            group_id INTEGER NOT NULL,
            FOREIGN KEY (group_id) REFERENCES groups (id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            date DATE NOT NULL,
            time TIME NOT NULL,
            FOREIGN KEY (group_id) REFERENCES groups (id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            student_id INTEGER,
            session_id INTEGER,
            status TEXT, -- 'present', 'absent', or 'justified'
            observation TEXT,
            PRIMARY KEY (student_id, session_id),
            FOREIGN KEY (student_id) REFERENCES students(id),
            FOREIGN KEY (session_id) REFERENCES sessions(id)
        )
    ''')
    conn.commit()
    conn.close()

# Initialize the database when the app starts
initialize_database()


# Route for the login page
@app.route('/')
def login():
    return render_template('login-page.html')


# Route to handle form submission (login)
@app.route('/submit', methods=['POST'])
def submit():
    fullname = request.form.get('fullname')
    password = request.form.get('password')

    # Check if the provided credentials match the hardcoded ones
    if fullname == "gl" and password == "gl":
        return redirect('/welcome')  # Render the welcome page
    else:
        return "Error: Invalid credentials, please try again."  # Error message if credentials don't match

   
@app.route('/welcome')
def welcome():
    conn = get_db_connection()
    try:
        # Fetch total number of classes
        total_classes = conn.execute('SELECT COUNT(*) FROM classes').fetchone()[0]

        # Fetch total number of groups
        total_groups = conn.execute('SELECT COUNT(*) FROM groups').fetchone()[0]

        # Fetch total number of students
        total_students = conn.execute('SELECT COUNT(*) FROM students').fetchone()[0]

        return render_template(
            'index.html',  # Your welcome page template
            total_classes=total_classes,
            total_groups=total_groups,
            total_students=total_students
        )
    except Exception as e:
        return f"An error occurred: {e}"
    finally:
        conn.close()




# Route to display all classes
@app.route('/classes')
def classes():
    conn = get_db_connection()
    classes = conn.execute('SELECT * FROM classes').fetchall()
    conn.close()
    return render_template('classes.html', classes=classes)

# Function to validate the year range format
def validate_year_range(year):
    # Use a regular expression to check the format
    if re.match(r'^\d{4}-\d{4}$', year):
        return True
    return False

# Route to add a class
@app.route('/add-class', methods=['GET', 'POST'])
def add_class():
    if request.method == 'POST':
        name = request.form['name']
        specialty = request.form['specialty']
        level = request.form['level']
        year = request.form['year']

        # Validate the year range
        if not validate_year_range(year):
            flash('Invalid year range format. Please use the format YYYY-YYYY (e.g., 2024-2025).', 'error')
            return redirect(url_for('add_class'))

        # If validation passes, insert into the database
        conn = get_db_connection()
        conn.execute(
            'INSERT INTO classes (name, specialty, level, year) VALUES (?, ?, ?, ?)',
            (name, specialty, level, year)
        )
        conn.commit()
        conn.close()
        flash('Class added successfully!', 'success')
        return redirect(url_for('classes'))

    return render_template('add-class.html')

# Route to edit a class
@app.route('/edit-class/<int:class_id>', methods=['GET', 'POST'])
def edit_class(class_id):
    conn = get_db_connection()
    class_data = conn.execute('SELECT * FROM classes WHERE id = ?', (class_id,)).fetchone()

    if request.method == 'POST':
        name = request.form['name']
        specialty = request.form['specialty']
        level = request.form['level']
        year = request.form['year']

        # Validate the year range
        if not validate_year_range(year):
            flash('Invalid year range format. Please use the format YYYY-YYYY (e.g., 2024-2025).', 'error')
            return redirect(url_for('edit_class', class_id=class_id))

        # If validation passes, update the database
        conn.execute(
            'UPDATE classes SET name = ?, specialty = ?, level = ?, year = ? WHERE id = ?',
            (name, specialty, level, year, class_id)
        )
        conn.commit()
        conn.close()
        flash('Class updated successfully!', 'success')
        return redirect(url_for('classes'))

    conn.close()
    return render_template('edit-class.html', class_data=class_data)
# Delete a class
@app.route('/delete-class/<int:class_id>', methods=['POST'])
def delete_class(class_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM classes WHERE id = ?', (class_id,))
    conn.commit()
    conn.close()
    return redirect('/classes')

# Display groups for a specific class
@app.route('/class/<int:class_id>/groups')
def groups(class_id):
    conn = get_db_connection()
    class_data = conn.execute('SELECT * FROM classes WHERE id = ?', (class_id,)).fetchone()
    groups = conn.execute('SELECT * FROM groups WHERE class_id = ?', (class_id,)).fetchall()
    conn.close()
    return render_template('groups.html', class_data=class_data, groups=groups, class_id=class_id)

# Add a group to a class
@app.route('/class/<int:class_id>/add-group', methods=['GET', 'POST'])
def add_group(class_id):
    if request.method == 'POST':

        group_type = request.form['group-type']
        if group_type =="TP/TD":
            conn = get_db_connection()
            conn.execute(
               'INSERT INTO groups ( type, class_id) VALUES (?, ?)',
               ( "TD", class_id)
               )
            conn.execute(
              'INSERT INTO groups ( type, class_id) VALUES (?, ?)',
              ( "TP", class_id)
              )
        else:    
             conn = get_db_connection()
             conn.execute(
             'INSERT INTO groups ( type, class_id) VALUES ( ?, ?)',
             ( group_type, class_id)
              )
        conn.commit()
        conn.close()
        return redirect(url_for('groups', class_id=class_id))

    return render_template('add-group.html', class_id=class_id)

# Edit a group
@app.route('/edit-group/<int:group_id>', methods=['GET', 'POST'])
def edit_group(group_id):
    conn = get_db_connection()
    group_data = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()

    if request.method == 'POST':
        group_type = request.form['group-type']

        conn.execute(
            'UPDATE groups SET type = ? WHERE id = ?',
            ( group_type, group_id)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('groups', class_id=group_data['class_id']))

    conn.close()
    return render_template('edit-group.html', group_data=group_data)

# Delete a group
@app.route('/delete-group/<int:group_id>', methods=['POST'])
def delete_group(group_id):
    conn = get_db_connection()
    group_data = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()

    conn.execute('DELETE FROM groups WHERE id = ?', (group_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('groups', class_id=group_data['class_id']))

@app.route('/export-attendance/<int:group_id>', methods=['GET', 'POST'])
def export_attendance(group_id):
    if request.method == 'POST':
        date_debut = request.form.get('date_debut')
        date_fin = request.form.get('date_fin')

        if not date_debut or not date_fin:
            return "رجاء تحقق من تاريخ .", 400

        conn = get_db_connection()

        sessions = conn.execute('''
            SELECT id, date 
            FROM sessions 
            WHERE group_id = ? AND date BETWEEN ? AND ?
            ORDER BY date
        ''', (group_id, date_debut, date_fin)).fetchall()

        if not sessions:
            conn.close()
            return "No sessions or wrong date.", 404

        students = conn.execute('''
            SELECT id, name, surname 
            FROM students 
            WHERE group_id = ?
            ORDER BY id
        ''', (group_id,)).fetchall()

        session_ids = tuple(session['id'] for session in sessions)
        attendance_data = conn.execute(f'''
            SELECT 
                attendance.student_id,
                attendance.session_id,
                attendance.status
            FROM attendance
            WHERE attendance.session_id IN ({','.join('?' * len(session_ids))})
        ''', session_ids).fetchall()
        conn.close()

        attendance_dict = {}
        for record in attendance_data:
            student_id = record['student_id']
            session_id = record['session_id']
            attendance_dict[(student_id, session_id)] = record['status']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Group {group_id} Attendance"

        headers = ['Student Name', 'Student Surname']
        headers.extend([f" ({session['date']})" for session in sessions])
        ws.append(headers)

        for student in students:
            row = [student['name'], student['surname']]
            for session in sessions:
                status = attendance_dict.get((student['id'], session['id']), 'Present') 
                row.append(status)
            ws.append(row)
        filepath = os.path.join(os.getcwd(), f"group_{group_id}_attendance_report.xlsx")       
        wb.save(filepath)

        return send_file(filepath, as_attachment=True, download_name=f'group_{group_id}_attendance_report.xlsx')

    return render_template('export-attendance.html')

@app.route('/group/<int:group_id>/students')
def view_students(group_id):
    try:
        conn = get_db_connection()

        # Fetch group details
        group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()
        if not group:
            conn.close()
            return f"Group with ID {group_id} not found.", 404

        # Fetch students in the group
        students = conn.execute('''
            SELECT s.id, s.name, s.surname, 
                   (SELECT COUNT(*) FROM attendance WHERE attendance.student_id = s.id AND attendance.status = 'present') AS sessions_attended
            FROM students s
            WHERE s.group_id = ?
        ''', (group_id,)).fetchall()

        # Get the total number of students
        total_students = len(students)

        conn.close()

        # Render the template with the students, group, and total_students
        return render_template('students.html', group=group, students=students, total_students=total_students)
    except Exception as e:
        return f"An error occurred: {str(e)}"

@app.route('/group/<int:group_id>/student/new', methods=['GET', 'POST'])
def add_student(group_id):
    conn = get_db_connection()
    group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()

    if not group:
        conn.close()
        return f"Group with ID {group_id} not found.", 404

    if request.method == 'POST':
        # Extract form data
        name = request.form.get('name')
        surname = request.form.get('surname')

        # Validate input
        if not name or not surname:
            conn.close()
            return "Name and surname are required.", 400

        # Insert new student into the database
        conn.execute(
            'INSERT INTO students (name, surname, group_id) VALUES (?, ?, ?)',
            (name, surname, group_id)
        )
        conn.commit()
        conn.close()

        # Redirect to the students view page
        return redirect(url_for('view_students', group_id=group_id))

    conn.close()
    # Render the add-student form
    return render_template('add-student.html', group=group)
@app.route('/add_students_excel/<int:group_id>', methods=['POST'])
def add_students_excel(group_id):
    if 'excel_file' not in request.files:
        return "No file part", 400

    file = request.files['excel_file']
    if file.filename == '':
        return "No selected file", 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return "Invalid file format. Please upload an Excel file.", 400

    # Load the Excel file
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    # Parse the data and insert into the database
    uploaded_students = []
    with get_db_connection() as conn:
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            name, surname = row
            if name and surname:  # Ensure both fields are provided
                conn.execute(
                    'INSERT INTO students (name, surname, group_id) VALUES (?, ?, ?)',
                    (name, surname, group_id)
                )
                uploaded_students.append({'name': name, 'surname': surname})
        conn.commit()

    # Fetch the group and render the template
    with get_db_connection() as conn:
        group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()
    return render_template('add-student.html', group=group, uploaded_students=uploaded_students)



@app.route('/export_students/<int:group_id>')
def export_students(group_id):
    # Fetch students from the database
    conn = get_db_connection()
    students = conn.execute(
        'SELECT name, surname FROM students WHERE group_id = ?', (group_id,)
    ).fetchall()
    conn.close()

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    # Add headers
    sheet.append(["Name", "Surname"])

    # Add student data
    for student in students:
        sheet.append([student["name"], student["surname"]])

    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Send the file as a response
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename=students_group_{group_id}.xlsx"}
    )

@app.route('/edit_student/<int:student_id>/<int:group_id>', methods=['GET', 'POST'])
def edit_student(student_id, group_id):
    conn = get_db_connection()

    # Fetch the student details
    student = conn.execute(
        'SELECT * FROM students WHERE id = ?', (student_id,)
    ).fetchone()

    if not student:
        conn.close()
        flash('Student not found.', 'error')
        return redirect(url_for('view_students', group_id=group_id))
    
    if request.method == 'POST':
        # Get the updated data from the form
        name = request.form.get('name')
        surname = request.form.get('surname')

        if not name or not surname:
            flash('Name and surname are required.', 'error')
        else:
            try:
                # Update the student's details in the database
                conn.execute(
                    'UPDATE students SET name = ?, surname = ? WHERE id = ?',
                    (name, surname, student_id)
                )
                conn.commit()
                flash('Student updated successfully!', 'success')
            except Exception as e:
                flash(f'Error updating student: {str(e)}', 'error')
        
        conn.close()
        return redirect(url_for('view_students', group_id=group_id))
    
    conn.close()
    # Render the edit form
    return render_template('edit-student.html', student=student, group_id=group_id)
@app.route('/group/<int:group_id>/student/<int:student_id>/delete', methods=['GET'])
def delete_student(group_id, student_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM students WHERE id = ?', (student_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('view_students', group_id=group_id))
@app.route('/group/<int:group_id>/student/delete', methods=['GET'])
def delete_students(group_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM students WHERE group_id = ?', (group_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('view_students', group_id=group_id))

@app.route('/session/<int:session_id>/manage_students', methods=['GET', 'POST'])
def manage_students(session_id):
    conn = get_db_connection()

    # Fetch session details
    session_data = conn.execute(
        'SELECT id, date, time FROM sessions WHERE id = ?', (session_id,)
    ).fetchone()

    if not session_data:
        conn.close()
        return "Session not found", 404

    # Fetch students in the group associated with this session
    group_id = conn.execute(
        'SELECT group_id FROM sessions WHERE id = ?', (session_id,)
    ).fetchone()['group_id']

    students = conn.execute(
        '''
        SELECT s.id, s.name, s.surname, a.status, a.observation
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id AND a.session_id = ?
        WHERE s.group_id = ?
        ''', (session_id, group_id)
    ).fetchall()

    conn.close()

    return render_template('manage_student.html', group_id=group_id,session_id=session_id, students=students)



@app.route('/save_attendance/<group_id>/<session_id>', methods=['POST'])
def save_attendance(group_id, session_id):
    conn = get_db_connection()
    
    try:
        # Get the list of students in the group
        students = conn.execute('SELECT * FROM students WHERE group_id = ?', (group_id,)).fetchall()

        for student in students:
            student_id = student['id']
            status = request.form.get(f'attendance_{student_id}[status]')
            observation = request.form.get(f'attendance_{student_id}[observation]')

            # Check if attendance record already exists for this student and session
            existing_attendance = conn.execute('SELECT * FROM attendance WHERE student_id = ? AND session_id = ?',
                                               (student_id, session_id)).fetchone()

            if existing_attendance:
                # Update the existing attendance record
                conn.execute('''UPDATE attendance 
                                SET status = ?, observation = ? 
                                WHERE student_id = ? AND session_id = ?''',
                             (status, observation, student_id, session_id))
            else:
                # Insert a new attendance record
                conn.execute('''INSERT INTO attendance (student_id, session_id, status, observation) 
                                VALUES (?, ?, ?, ?)''',
                             (student_id, session_id, status, observation))

            # If student is present, update the number of sessions attended
            if status == 'present':
                conn.execute('UPDATE students SET sessions_attended = sessions_attended + 1 WHERE id = ?',
                             (student_id,))

        # Commit the changes
        conn.commit()

        return redirect(url_for('view_sessions', group_id=group_id))
    except Exception as e:
        conn.rollback()
        return f"An error occurred: {str(e)}", 500
    finally:
        conn.close()


@app.route('/group/<int:group_id>/sessions', methods=['GET'])
def view_sessions(group_id):
    conn = get_db_connection()
    try:
        # Fetch all sessions for the given group
        sessions = conn.execute('SELECT * FROM sessions WHERE group_id = ?', (group_id,)).fetchall()

        # Fetch the class_id for navigation
        group_info = conn.execute('SELECT class_id FROM groups WHERE id = ?', (group_id,)).fetchone()
        if not group_info:
            return "Group not found", 404

        class_id = group_info['class_id']

        # For each session, count the number of present and absent students
        session_data = []
        for session in sessions:
            session_id = session['id']

            # Count present students
            present_count = conn.execute('''
                SELECT COUNT(*) 
                FROM attendance 
                WHERE session_id = ? AND status = 'present'
            ''', (session_id,)).fetchone()[0]

            # Count absent students
            absent_count = conn.execute('''
                SELECT COUNT(*) 
                FROM attendance 
                WHERE session_id = ? AND status = 'absent'
            ''', (session_id,)).fetchone()[0]

            # Add session details with counts to the session_data list
            session_data.append({
                'id': session['id'],
                'date': session['date'],
                'time': session['time'],
                'present_count': present_count,
                'absent_count': absent_count
            })

        return render_template('sessions.html', sessions=session_data, group_id=group_id, class_id=class_id)

    except Exception as e:
        return f"An error occurred: {e}"

    finally:
        conn.close()

@app.route('/group/<int:group_id>/session/add', methods=['GET', 'POST'])
def add_session(group_id):
    conn = get_db_connection()
    group = conn.execute('SELECT * FROM groups WHERE id = ?', (group_id,)).fetchone()
    if not group:
        return "Group not found", 404

    # Pass class_id to the template
    class_id = group['class_id']

    if request.method == 'POST':
        session_date = request.form['session-date']
        session_time = request.form['session-time']

        conn.execute(
            'INSERT INTO sessions (group_id, date, time) VALUES (?, ?, ?)',
            (group_id, session_date, session_time)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('view_sessions', group_id=group_id))

    conn.close()
    return render_template('add-session.html', group_id=group_id, class_id=class_id)

@app.route('/group/<int:group_id>/session/edit/<int:session_id>', methods=['GET', 'POST'])
def edit_session(group_id, session_id):
    conn = get_db_connection()
    session = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()

    if not session:
        conn.close()
        return "Session not found", 404

    if request.method == 'POST':
        session_date = request.form['session-date']
        session_time = request.form['session-time']

        try:
            conn.execute(
                'UPDATE sessions SET date = ?, time = ? WHERE id = ?',
                (session_date, session_time, session_id)
            )
            conn.commit()
            return redirect(url_for('view_sessions', group_id=group_id))
        except Exception as e:
            return f"An error occurred while updating the session: {e}"
        finally:
            conn.close()

    conn.close()
    return render_template('edit-session.html', session=session, group_id=group_id)

@app.route('/group/<int:group_id>/session/delete/<int:session_id>', methods=['POST'])
def delete_session(group_id, session_id):
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM sessions WHERE id = ?', (session_id,))
        conn.commit()
        return redirect(url_for('view_sessions', group_id=group_id))
    except Exception as e:
        return f"An error occurred while deleting the session: {e}"
    finally:
        conn.close()



@app.route('/export_session/<int:session_id>')
def export_session(session_id):
    conn = get_db_connection()

    # Retrieve session details
    session = conn.execute('SELECT * FROM sessions WHERE id = ?', (session_id,)).fetchone()
    if not session:
        return "Session not found", 404

    # Retrieve student attendance for the session
    students = conn.execute('''
        SELECT s.name, s.surname, a.status, a.observation
        FROM students s
        JOIN attendance a ON s.id = a.student_id
        WHERE a.session_id = ?
    ''', (session_id,)).fetchall()

    conn.close()

    # Create Excel file in memory
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Session Attendance')

    # Write session details
    worksheet.write('A1', 'Session ID')
    worksheet.write('B1', session['id'])
    worksheet.write('A2', 'Group ID')
    worksheet.write('B2', session['group_id'] if 'group_id' in session.keys() else 'N/A')
    worksheet.write('A3', 'Name')
    worksheet.write('B3', session['name'] if 'name' in session.keys() else 'N/A')
    worksheet.write('A4', 'Date')
    worksheet.write('B4', session['date'] if 'date' in session.keys() else 'N/A')
    worksheet.write('A5', 'Time')
    worksheet.write('B5', session['time'] if 'time' in session.keys() else 'N/A')

    # Write attendance table header
    worksheet.write('A7', 'Name')
    worksheet.write('B7', 'Surname')
    worksheet.write('C7', 'Status')
    worksheet.write('D7', 'Observation')

    # Write attendance data
    row = 7
    for student in students:
        worksheet.write(row, 0, student['name'])
        worksheet.write(row, 1, student['surname'])
        worksheet.write(row, 2, student['status'])
        worksheet.write(row, 3, student['observation'])
        row += 1 

    workbook.close()
    output.seek(0)

    # Send the file as a response
    return send_file(
        output,
        as_attachment=True,
        download_name=f'session_{session_id}_attendance.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# Main route
@app.route('/')
def index():
    return redirect('/classes')


if __name__ == '__main__':
    root = tk.Tk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.destroy()

    Window = webview.create_window('Management Student', app , width=screen_width, height=screen_height, resizable=True)

    def toggle_fullscreen():
        Window.toggle_fullscreen()
        
        keyboard.add_hotkey("$", toggle_fullscreen)

    # Start the WebView
    webview.start()
    app.run(debug=True)
